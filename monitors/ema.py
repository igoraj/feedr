import requests
import openpyxl
import re
import xml.etree.ElementTree as ET
import os
import json
from typing import List, Dict, Any

from monitors.base import BaseMonitor

# Suppress SSL warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class EmaMonitor(BaseMonitor):
    def __init__(self, substances: List[str]):
        self.substances = substances
        self.ema_base_url = "https://www.ema.europa.eu"
        self.xlsx_url = f"{self.ema_base_url}/en/documents/report/medicines-output-medicines-report_en.xlsx"
        self.xlsx_filename = "medicines_report.xlsx"
        self.state_filename = "ema_monitor_state.json"

    @property
    def name(self) -> str:
        return "EMA"

    def _download_file(self) -> bool:
        print(f"Downloading medicines database from {self.xlsx_url}...")
        try:
            response = requests.get(self.xlsx_url, verify=False, headers={'User-Agent': 'Mozilla/5.0'})
            response.raise_for_status()
            with open(self.xlsx_filename, 'wb') as f:
                f.write(response.content)
            print("Download complete.")
            return True
        except Exception as e:
            print(f"Error downloading file: {e}")
            return False

    def _find_medicines_by_substances(self) -> Dict[str, List[Dict[str, str]]]:
        print(f"Searching for substances {self.substances} in database...")
        workbook = openpyxl.load_workbook(self.xlsx_filename, read_only=True)
        sheet = workbook.active

        headers = {}
        header_row_idx = -1
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_values = [str(cell).lower() for cell in row if cell]
            if any("active substance" in val for val in row_values):
                header_row_idx = row_idx
                for col_idx, cell in enumerate(row):
                    if cell:
                        headers[str(cell).strip()] = col_idx
                break

        if header_row_idx == -1:
            print("Could not find header row.")
            return {}

        name_col, substance_col, url_col = -1, -1, -1
        for h, idx in headers.items():
            h_lower = h.lower()
            if "name" in h_lower and "medicine" in h_lower: name_col = idx
            elif "active substance" in h_lower: substance_col = idx
            elif "url" in h_lower and "medicine" in h_lower: url_col = idx

        if name_col == -1 or substance_col == -1 or url_col == -1:
            print("Could not identify necessary columns.")
            return {}

        results = {sub: [] for sub in self.substances}
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or len(row) <= max(name_col, substance_col, url_col): continue
            prod_substance = str(row[substance_col]).lower() if row[substance_col] else ""
            for sub in self.substances:
                if sub.lower() in prod_substance:
                    results[sub].append({"name": row[name_col], "url": row[url_col]})
        return results

    def _get_latest_rss_update(self, product_url):
        try:
            response = requests.get(product_url, verify=False, headers={'User-Agent': 'Mozilla/5.0'})
            if response.status_code != 200: return "Error: Page not reachable", None, None, None
            html = response.text

            # More robust regex to find the RSS link
            rss_match = re.search(r'<a[^>]+href="([^"]+)"[^>]+class="[^"]*ema-rss-button[^"]*"', html, re.IGNORECASE)
            if not rss_match:
                rss_match = re.search(r'<a[^>]+class="[^"]*ema-rss-button[^"]*"[^>]+href="([^"]+)"', html, re.IGNORECASE)

            if rss_match:
                rss_path = rss_match.group(1)
                rss_url = f"{self.ema_base_url}{rss_path}" if rss_path.startswith("/") else rss_path
                rss_response = requests.get(rss_url, verify=False, headers={'User-Agent': 'Mozilla/5.0'})
                root = ET.fromstring(rss_response.content)
                item = root.find(".//item")
                if item is not None:
                    title = item.find("title").text.strip() if item.find("title") is not None else "No Title"
                    pub_date = item.find("pubDate").text.strip() if item.find("pubDate") is not None else "No Date"
                    link = item.find("link").text.strip() if item.find("link") is not None else "No Link"
                    description = item.find("description").text.strip() if item.find("description") is not None else ""
                    return title, pub_date, link, description
            return "No RSS found", None, None, None
        except Exception as e:
            return f"Error: {str(e)}", None, None, None

    def _load_state(self):
        if os.path.exists(self.state_filename):
            try:
                with open(self.state_filename, 'r') as f:
                    return json.load(f)
            except json.JSONDecodeError:
                return {}
        return {}

    def _save_state(self, state):
        with open(self.state_filename, 'w') as f:
            json.dump(state, f, indent=4)

    def fetch_items(self) -> (List[Dict[str, Any]], List[Dict[str, Any]]):
        if not self._download_file():
            return [], []

        findings = self._find_medicines_by_substances()
        state = self._load_state()
        all_items = []
        new_updates = []

        print("\nChecking for updates...")
        for substance, products in findings.items():
            for prod in products:
                prod_name = prod['name']
                if not prod['url']: continue

                title, date, link, description = self._get_latest_rss_update(prod['url'])
                if not date or date == "No Date": continue

                item = {
                    "substance": substance,
                    "product": prod_name,
                    "date": date,
                    "title": title,
                    "link": link,
                    "description": description or f"Update for {prod_name}",
                    "id": link
                }
                all_items.append(item)

                update_key = f"{date}_{title}"
                if prod_name not in state or state[prod_name] != update_key:
                    new_updates.append(item)
                    state[prod_name] = update_key

        self._save_state(state)
        if os.path.exists(self.xlsx_filename):
            os.remove(self.xlsx_filename)

        return all_items, new_updates

    def check_for_updates(self) -> List[Dict[str, Any]]:
        # This method is now a simplified wrapper for fetch_items for backward compatibility
        # or for cases where only new updates are needed.
        _, new_updates = self.fetch_items()
        return new_updates
